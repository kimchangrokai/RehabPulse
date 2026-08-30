"""엑셀 업서트 — openpyxl 기반 단일 워크북 상태 저장소.

시트: 사건목록, 일반내용, 진행명령, 변경이력, 실행로그, 종료목록
"""

from __future__ import annotations

import hashlib
import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from ..models import GeneralContent, OrderRow, CaseRecord, ChangeEvent

logger = logging.getLogger(__name__)

# ── 시트 헤더 정의 ──────────────────────────────────────────────────

CASE_LIST_HEADERS = [
    "법원", "사건번호", "년도", "사건구분", "일련번호", "당사자명",
    "활성", "인가여부", "consecutive_miss_days",
    "최근조회결과", "최근오류", "최종확인일시",
    "최초수집일시", "최종변경일시",
]

GENERAL_HEADERS = [
    "법원", "사건번호", "당사자명",
    "사건명", "재판부/회생위원", "접수일",
    "개시결정일", "변제계획인가일", "면책결정일", "절차폐지결정일", "종국결과",
    "최초수집일시", "최종확인일시", "최종변경일시", "row_hash",
]

ORDER_HEADERS = [
    "법원", "사건번호", "당사자명",
    "일자", "진행구분", "내용", "결과",
    "삭제됨",
    "최초수집일시", "최종확인일시", "최종변경일시", "row_hash",
]

HISTORY_HEADERS = [
    "실행시각", "법원", "사건번호", "당사자명", "이벤트", "상세",
]

RUNLOG_HEADERS = [
    "실행시각", "대상수", "성공", "실패", "결번", "비고",
]

ARCHIVE_HEADERS = [
    "법원", "사건번호", "당사자명", "종료일", "사유",
]

SHEET_DEFS = {
    "사건목록": CASE_LIST_HEADERS,
    "일반내용": GENERAL_HEADERS,
    "진행명령": ORDER_HEADERS,
    "변경이력": HISTORY_HEADERS,
    "실행로그": RUNLOG_HEADERS,
    "종료목록": ARCHIVE_HEADERS,
}


class ExcelStore:
    """엑셀 워크북 상태 저장소."""

    def __init__(self, path: str | Path, backup_dir: str | Path = "backup",
                 retention: int = 30):
        self.path = Path(path)
        self.backup_dir = Path(backup_dir)
        self.retention = retention
        self.wb: Optional[openpyxl.Workbook] = None

    # ── 로드 / 저장 ──────────────────────────────────────────────

    def load(self) -> None:
        """기존 워크북을 열거나 새로 만든다. 시트·헤더 보장."""
        if self.path.exists():
            self.wb = openpyxl.load_workbook(self.path)
        else:
            self.wb = openpyxl.Workbook()
            # 기본 Sheet 제거
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]

        for sheet_name, headers in SHEET_DEFS.items():
            if sheet_name not in self.wb.sheetnames:
                ws = self.wb.create_sheet(title=sheet_name)
                ws.append(headers)
            else:
                ws = self.wb[sheet_name]
                # 누락된 헤더 열 추가
                existing = [cell.value for cell in ws[1]]
                for h in headers:
                    if h not in existing:
                        ws.cell(row=1, column=len(existing) + 1, value=h)
                        existing.append(h)

    def save(self) -> None:
        """워크북 저장. 잠금 시 임시 파일로 안내."""
        if self.wb is None:
            return
        self._backup()
        try:
            self.wb.save(self.path)
            logger.info(f"엑셀 저장 완료: {self.path}")
        except PermissionError:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            tmp = self.path.with_name(f"{self.path.stem}_{ts}.xlsx")
            self.wb.save(tmp)
            logger.warning(f"엑셀 잠금! 임시 파일 저장: {tmp}")

    def _backup(self) -> None:
        """저장 전 백업. retention 개수만큼 오래된 백업 정리."""
        if not self.path.exists():
            return
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dst = self.backup_dir / f"{self.path.stem}_{ts}.xlsx"
        shutil.copy2(self.path, dst)
        # 오래된 백업 정리
        backups = sorted(self.backup_dir.glob(f"{self.path.stem}_*.xlsx"))
        while len(backups) > self.retention:
            oldest = backups.pop(0)
            oldest.unlink(missing_ok=True)

    # ── 사건목록 관리 ────────────────────────────────────────────

    def add_case(self, record: CaseRecord) -> bool:
        """사건을 등록한다. 이미 있으면 False."""
        ws = self.wb["사건목록"]
        key = (record.court, record.case_no)
        for row in ws.iter_rows(min_row=2, values_only=False):
            vals = [c.value for c in row]
            if (vals[0], vals[1]) == key:
                return False  # 이미 존재

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([
            record.court, record.case_no, record.year,
            record.case_type, record.serial, record.party,
            record.active, record.plan_approved,
            record.consecutive_miss_days,
            record.last_result, record.last_error, "",
            now, "",
        ])
        return True

    def remove_case(self, court: str, case_no: str, purge: bool = False) -> bool:
        """사건을 비활성화(기본) 또는 삭제(purge=True)한다."""
        ws = self.wb["사건목록"]
        for row_idx in range(2, ws.max_row + 1):
            if (ws.cell(row_idx, 1).value == court and
                    ws.cell(row_idx, 2).value == case_no):
                if purge:
                    ws.delete_rows(row_idx)
                else:
                    ws.cell(row_idx, 7).value = "N"  # 활성=N
                return True
        return False

    def get_active_cases(self) -> list[CaseRecord]:
        """활성 사건 목록을 반환한다."""
        ws = self.wb["사건목록"]
        cases = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            if row[6] != "Y":  # 활성 != Y
                continue
            cases.append(CaseRecord(
                court=row[0], case_no=row[1],
                year=row[2] or "", case_type=row[3] or "",
                serial=row[4] or "", party=row[5] or "",
                active=row[6] or "Y",
                plan_approved=row[7] or "",
                consecutive_miss_days=int(row[8] or 0),
                last_result=row[9] or "",
                last_error=row[10] or "",
            ))
        return cases

    def update_case_status(self, court: str, case_no: str, **kwargs) -> None:
        """사건목록의 특정 필드를 업데이트한다.

        지원 키: active, plan_approved, consecutive_miss_days,
                 last_result, last_error, last_check
        """
        ws = self.wb["사건목록"]
        col_map = {
            "active": 7, "plan_approved": 8, "consecutive_miss_days": 9,
            "last_result": 10, "last_error": 11, "last_check": 12,
        }
        for row_idx in range(2, ws.max_row + 1):
            if (ws.cell(row_idx, 1).value == court and
                    ws.cell(row_idx, 2).value == case_no):
                for key, val in kwargs.items():
                    col = col_map.get(key)
                    if col:
                        ws.cell(row_idx, col).value = val
                return

    # ── 일반내용 업서트 ──────────────────────────────────────────

    def upsert_general(self, general: GeneralContent, party: str) -> dict:
        """일반내용을 업서트한다. 변경 정보 반환."""
        ws = self.wb["일반내용"]
        key = (general.court, general.case_no)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row_hash = _hash_general(general)

        # 기존 행 찾기
        for row_idx in range(2, ws.max_row + 1):
            if (ws.cell(row_idx, 1).value == general.court and
                    ws.cell(row_idx, 2).value == general.case_no):
                old_hash = ws.cell(row_idx, 15).value  # row_hash 열
                if old_hash == row_hash:
                    # 동일 — 최종확인일시만 갱신
                    ws.cell(row_idx, 13).value = now
                    return {"changed": False}
                # 변경 — 값 업데이트
                _write_general_row(ws, row_idx, general, party, now, row_hash)
                return {"changed": True, "old_hash": old_hash}

        # 신규 행
        row_idx = ws.max_row + 1
        _write_general_row(ws, row_idx, general, party, now, row_hash)
        ws.cell(row_idx, 12).value = now  # 최초수집일시
        return {"changed": True, "old_hash": None}

    def read_general(self, court: str, case_no: str) -> Optional[dict]:
        """이전 일반내용 스냅샷을 읽는다."""
        ws = self.wb["일반내용"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == court and row[1] == case_no:
                return {
                    "court": row[0], "case_no": row[1], "party": row[2],
                    "case_name": row[3], "panel": row[4], "filed_date": row[5],
                    "commencement_date": row[6], "plan_approved_date": row[7],
                    "discharge_date": row[8], "revocation_date": row[9],
                    "terminal_result": row[10],
                }
        return None

    # ── 진행명령 업서트 ──────────────────────────────────────────

    def upsert_orders(self, orders: list[OrderRow], party: str,
                      court: str = "", case_no: str = "") -> dict:
        """진행명령을 사건 단위로 업서트한다. 다른 사건 행은 건드리지 않는다."""
        ws = self.wb["진행명령"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if orders:
            court = court or orders[0].court
            case_no = case_no or orders[0].case_no
        if not court or not case_no:
            return {"added": 0, "updated": 0, "removed": 0}

        # 기존 행 인덱스 — 이 사건만 (키: 일자+내용)
        existing = {}
        for row_idx in range(2, ws.max_row + 1):
            c = ws.cell(row_idx, 1).value
            cn = ws.cell(row_idx, 2).value
            if c != court or cn != case_no:
                continue
            d = ws.cell(row_idx, 4).value
            ct = ws.cell(row_idx, 6).value
            existing[(c, cn, d, ct)] = row_idx

        added = updated = removed = 0
        seen_keys = set()

        for order in orders:
            key = (order.court, order.case_no, order.date, order.content)
            seen_keys.add(key)
            row_hash = _hash_order(order)

            if key in existing:
                row_idx = existing[key]
                old_hash = ws.cell(row_idx, 12).value
                if old_hash != row_hash:
                    # 내용/결과 변경
                    _write_order_row(ws, row_idx, order, party, now, row_hash)
                    ws.cell(row_idx, 8).value = ""  # 삭제됨 해제
                    updated += 1
                else:
                    ws.cell(row_idx, 11).value = now  # 최종확인일시
                del existing[key]
            else:
                # 신규 행
                row_idx = ws.max_row + 1
                _write_order_row(ws, row_idx, order, party, now, row_hash)
                ws.cell(row_idx, 9).value = now  # 최초수집일시
                added += 1

        # 사라진 행 → 삭제됨 플래그
        for key, row_idx in existing.items():
            if ws.cell(row_idx, 8).value != "Y":  # 아직 삭제됨이 아니면
                ws.cell(row_idx, 8).value = "Y"
                ws.cell(row_idx, 11).value = now
                removed += 1

        return {"added": added, "updated": updated, "removed": removed}

    def read_orders(self, court: str, case_no: str) -> list[dict]:
        """이전 진행명령 스냅샷을 읽는다."""
        ws = self.wb["진행명령"]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == court and row[1] == case_no:
                if row[7] == "Y":  # 삭제됨
                    continue
                rows.append({
                    "court": row[0], "case_no": row[1], "party": row[2],
                    "date": row[3], "category": row[4],
                    "content": row[5], "result": row[6],
                })
        return rows

    # ── 변경이력 / 실행로그 ──────────────────────────────────────

    def append_history(self, event: ChangeEvent) -> None:
        """변경이력에 이벤트를 추가한다."""
        ws = self.wb["변경이력"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([now, event.court, event.case_no, event.party,
                   event.event, event.detail])

    def append_runlog(self, total: int, success: int, fail: int,
                      miss: int, note: str = "") -> None:
        """실행로그에 기록한다."""
        ws = self.wb["실행로그"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([now, total, success, fail, miss, note])

    # ── 종료목록 ─────────────────────────────────────────────────

    def archive_case(self, court: str, case_no: str, party: str,
                     reason: str = "3일 연속 조회 없음") -> None:
        """종료목록으로 이동."""
        ws = self.wb["종료목록"]
        today = datetime.now().strftime("%Y-%m-%d")
        ws.append([court, case_no, party, today, reason])

    # ── 유틸 ─────────────────────────────────────────────────────

    def beautify(self) -> None:
        """셀 서식: 헤더 굵게, 자동 필터, 열 너비."""
        for sheet_name in SHEET_DEFS:
            ws = self.wb[sheet_name]
            if ws.max_row < 2:
                continue
            # 헤더 슀식
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            # 자동 필터
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            # 기본 열 너비
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                if ws.column_dimensions[col_letter].width is None:
                    ws.column_dimensions[col_letter].width = 15


# ── 내부 헬퍼 ────────────────────────────────────────────────────────

def _hash_general(g: GeneralContent) -> str:
    """일반내용의 해시."""
    parts = [
        g.case_name, g.panel, g.filed_date,
        g.commencement_date, g.plan_approved_date,
        g.discharge_date, g.revocation_date, g.terminal_result,
    ]
    return hashlib.md5("|".join(str(p) for p in parts).encode()).hexdigest()[:12]


def _hash_order(o: OrderRow) -> str:
    """명령 행의 해시."""
    parts = [o.date, o.category, o.content, o.result]
    return hashlib.md5("|".join(parts).encode()).hexdigest()[:12]


def _write_general_row(ws, row_idx: int, g: GeneralContent,
                       party: str, now: str, row_hash: str) -> None:
    """일반내용 행을 쓴다."""
    ws.cell(row_idx, 1, g.court)
    ws.cell(row_idx, 2, g.case_no)
    ws.cell(row_idx, 3, party)
    ws.cell(row_idx, 4, g.case_name)
    ws.cell(row_idx, 5, g.panel)
    ws.cell(row_idx, 6, g.filed_date)
    ws.cell(row_idx, 7, g.commencement_date)
    ws.cell(row_idx, 8, g.plan_approved_date)
    ws.cell(row_idx, 9, g.discharge_date)
    ws.cell(row_idx, 10, g.revocation_date)
    ws.cell(row_idx, 11, g.terminal_result)
    ws.cell(row_idx, 13, now)  # 최종확인일시
    ws.cell(row_idx, 14, now)  # 최종변경일시
    ws.cell(row_idx, 15, row_hash)


def _write_order_row(ws, row_idx: int, o: OrderRow,
                     party: str, now: str, row_hash: str) -> None:
    """명령 행을 쓴다."""
    ws.cell(row_idx, 1, o.court)
    ws.cell(row_idx, 2, o.case_no)
    ws.cell(row_idx, 3, party)
    ws.cell(row_idx, 4, o.date)
    ws.cell(row_idx, 5, o.category)
    ws.cell(row_idx, 6, o.content)
    ws.cell(row_idx, 7, o.result)
    ws.cell(row_idx, 10, now)  # 최종확인일시
    ws.cell(row_idx, 11, now)  # 최종변경일시
    ws.cell(row_idx, 12, row_hash)
